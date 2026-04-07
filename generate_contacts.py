"""
Generate partner_contacts.py from 3 Excel source files.
Reads commercial contacts, technical contacts, and feature/vertical data,
then writes a consolidated Python dictionary to data/partner_contacts.py.
"""

import os
import re
import openpyxl

# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------
SALES_FILE = "/Users/danielareyes/Downloads/Sales Contacts PARTNERS YUNO 2025.xlsx"
TECH_FILE = "/Users/danielareyes/Downloads/Partner Technical Contact List (1).xlsx"
SOT_FILE = "/Users/danielareyes/Downloads/[WIP] Source of Truth v.1 (2).xlsx"

OUTPUT_DIR = os.path.join(os.path.dirname(__file__), "data")
OUTPUT_FILE = os.path.join(OUTPUT_DIR, "partner_contacts.py")

YUNO_EXCLUDE_NAMES = {
    n.strip().lower()
    for n in [
        "Daniela Reyes",
        "Francisco Quintana",
        "Johanderson Guevara",
        "Talita Diaz Gama",
        "Alessandra Rospigliosi",
        "Jorge Restrepo",
        "Valentina Perez",
        "Lily",
        "Partha",
    ]
}

FEATURE_COLS = [
    "ACCEPTS_GAMBLING",
    "ACCEPTS_GAMING",
    "ACCEPTS_CRYPTO",
    "ACCEPTS_FOREX",
    "ACCEPTS_ADULT",
    "ACCEPTS_HIGH_RISK",
    "ACCEPTS_AIRLINES",
    "ACCEPTS_MULTI_LEVEL_MARKETING",
    "SUPPORTS_TOKENIZATION",
    "SUPPORTS_RECURRING_PAYMENTS",
    "SUPPORTS_PAYOUTS",
    "SUPPORTS_INSTALLMENTS",
    "SUPPORTS_NETWORK_TOKENS",
]

FEATURE_LABELS = {
    "ACCEPTS_GAMBLING": "Gambling",
    "ACCEPTS_GAMING": "Gaming",
    "ACCEPTS_CRYPTO": "Crypto",
    "ACCEPTS_FOREX": "Forex",
    "ACCEPTS_ADULT": "Adult",
    "ACCEPTS_HIGH_RISK": "High Risk",
    "ACCEPTS_AIRLINES": "Airlines",
    "ACCEPTS_MULTI_LEVEL_MARKETING": "Multi-Level Marketing",
    "SUPPORTS_TOKENIZATION": "Tokenization",
    "SUPPORTS_RECURRING_PAYMENTS": "Recurring Payments",
    "SUPPORTS_PAYOUTS": "Payouts",
    "SUPPORTS_INSTALLMENTS": "Installments",
    "SUPPORTS_NETWORK_TOKENS": "Network Tokens",
}

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def clean(val):
    """Return stripped string or empty string for None/NaN."""
    if val is None:
        return ""
    s = str(val).strip()
    if s.lower() in ("nan", "none", "n/a", "na"):
        return ""
    return s


def norm_partner(name):
    """Normalize partner name to uppercase, strip whitespace."""
    return clean(name).upper().strip().rstrip("/").strip()


def is_yuno_contact(name, email):
    """Check if a contact is an internal Yuno person to exclude."""
    n = clean(name).lower()
    e = clean(email).lower()
    if "@y.uno" in e or "@yuno" in e:
        return True
    for excl in YUNO_EXCLUDE_NAMES:
        if excl in n:
            return True
    return False


def extract_email(text):
    """Extract first email address from free-text."""
    text = clean(text)
    if not text:
        return ""
    emails = re.findall(r"[A-Za-z0-9._%+\-]+@[A-Za-z0-9.\-]+\.[A-Za-z]{2,}", text)
    # Filter out yuno emails
    for em in emails:
        if "@y.uno" not in em.lower() and "@yuno" not in em.lower():
            return em
    return emails[0] if emails else ""


def extract_name_from_text(text):
    """Try to extract a person's name from free-text technical contact field."""
    text = clean(text)
    if not text:
        return ""
    # Look for patterns like "Name - email" or "Name\nemail"
    # Common patterns: "Name - email@...", "Name: email@..."
    # Try to find name before an email
    lines = re.split(r"\n+", text)
    for line in lines:
        line = line.strip()
        if not line:
            continue
        # Skip lines that are just emails or support@
        if re.match(r"^[A-Za-z0-9._%+\-]+@", line):
            continue
        if line.lower().startswith(("email:", "phone:", "availability:", "scope:",
                                     "24/7", "for 24/7", "tel.", "e-mail")):
            continue
        # "Name - email@..."
        m = re.match(r"^(?:CC:\s*)?([A-Z][a-z]+(?:\s+[A-Z][a-z]+)+)\s*[-–]\s*\S+@", line)
        if m:
            return m.group(1).strip()
        # "Name email@..."
        m = re.match(r"^([A-Z][a-z]+(?:\s+[A-Z][a-z]+)+)\s+\S+@", line)
        if m:
            return m.group(1).strip()
        # "For new integrations contact - name@..."  -> skip
        if line.lower().startswith("for "):
            # "For new integrations contact - jeffrey.vanderlinden@worldline.com"
            # try to extract name from email
            continue
        # Line with just a name (2-3 capitalized words, no @ symbol)
        if "@" not in line and re.match(r"^[A-Z][a-z]+(?:\s+[A-Z][a-z]+){1,3}$", line.strip()):
            return line.strip()
    # Fallback: try to get name from email prefix
    email = extract_email(text)
    if email:
        prefix = email.split("@")[0]
        # If prefix looks like first.last or first_last
        parts = re.split(r"[._]", prefix)
        if len(parts) >= 2 and all(p.isalpha() for p in parts):
            return " ".join(p.capitalize() for p in parts)
    return ""


def is_truthy(val):
    """Check if a cell value represents True/1."""
    if val is True:
        return True
    if isinstance(val, (int, float)) and val == 1:
        return True
    s = clean(val).lower()
    return s in ("true", "1", "1.0", "yes")


def escape_str(s):
    """Escape a string for Python repr."""
    return s.replace("\\", "\\\\").replace('"', '\\"')


# ---------------------------------------------------------------------------
# 1. Read Sales Contacts
# ---------------------------------------------------------------------------
print("Reading Sales Contacts...")
contacts = {}  # partner_upper -> dict

wb = openpyxl.load_workbook(SALES_FILE, read_only=True, data_only=True)

# Sheet: "Partnerships Contacts"
ws = wb["Partnerships Contacts"]
rows = list(ws.iter_rows(values_only=True))
header = [clean(h) for h in rows[0]]

col_partner = header.index("Partners")
col_contact = header.index("Partnerships AM")  # actual contact name column
col_email = header.index("AM Email")
col_phone = header.index("AM Phone")
col_role = header.index("AM Role")

for row in rows[1:]:
    partner = norm_partner(row[col_partner])
    if not partner:
        continue
    name = clean(row[col_contact])
    email = clean(row[col_email])
    phone = clean(row[col_phone])
    if is_yuno_contact(name, email):
        continue
    if partner not in contacts:
        contacts[partner] = {
            "commercial_contact": "",
            "commercial_email": "",
            "commercial_phone": "",
            "technical_contact": "",
            "technical_email": "",
            "technical_phone": "",
            "verticals": "",
        }
    # Only set if we have a non-empty contact and current is empty
    entry = contacts[partner]
    if name and not entry["commercial_contact"]:
        entry["commercial_contact"] = name
    if email and not entry["commercial_email"]:
        entry["commercial_email"] = email
    if phone and not entry["commercial_phone"]:
        entry["commercial_phone"] = phone

# Sheet: "All contacts partners" (header at row index 1)
ws2 = wb["All contacts partners"]
rows2 = list(ws2.iter_rows(values_only=True))
header2 = [clean(h) for h in rows2[1]]  # row index 1 is header

col_p2 = header2.index("Partner")
col_name2 = header2.index("Contact name")
col_email2 = header2.index("Email")
col_phone2 = header2.index("Phone number")
col_pos2 = header2.index("Position")

for row in rows2[2:]:
    partner = norm_partner(row[col_p2])
    if not partner:
        continue
    name = clean(row[col_name2])
    email = clean(row[col_email2])
    phone = clean(row[col_phone2])
    if is_yuno_contact(name, email):
        continue
    if partner not in contacts:
        contacts[partner] = {
            "commercial_contact": "",
            "commercial_email": "",
            "commercial_phone": "",
            "technical_contact": "",
            "technical_email": "",
            "technical_phone": "",
            "verticals": "",
        }
    entry = contacts[partner]
    if name and not entry["commercial_contact"]:
        entry["commercial_contact"] = name
    if email and not entry["commercial_email"]:
        entry["commercial_email"] = email
    if phone and not entry["commercial_phone"]:
        entry["commercial_phone"] = phone

wb.close()
print(f"  Found {len(contacts)} partners from Sales Contacts.")

# ---------------------------------------------------------------------------
# 2. Read Technical Contacts
# ---------------------------------------------------------------------------
print("Reading Technical Contacts...")

wb2 = openpyxl.load_workbook(TECH_FILE, read_only=True, data_only=True)
ws_tech = wb2["LATAM"]
rows_tech = list(ws_tech.iter_rows(values_only=True))

# Header is at row index 5 (0-indexed), which is rows_tech[5]
header_tech = [clean(h) for h in rows_tech[5]]

col_provider = header_tech.index("Provider")
col_day2day = header_tech.index("Technical Contact (Day to Day)")
col_p1 = header_tech.index("Technical Contact P1")

for row in rows_tech[6:]:
    if len(row) <= col_provider:
        continue
    provider = norm_partner(row[col_provider])
    if not provider:
        continue
    day2day = clean(row[col_day2day]) if len(row) > col_day2day else ""
    p1 = clean(row[col_p1]) if len(row) > col_p1 else ""

    # Prefer day-to-day contact
    tech_text = day2day or p1
    tech_name = extract_name_from_text(tech_text)
    tech_email = extract_email(tech_text)

    if is_yuno_contact(tech_name, tech_email):
        tech_name = ""
        tech_email = ""
        # Try P1 fallback
        if p1 and p1 != day2day:
            tech_name = extract_name_from_text(p1)
            tech_email = extract_email(p1)
            if is_yuno_contact(tech_name, tech_email):
                tech_name = ""
                tech_email = ""

    if provider not in contacts:
        contacts[provider] = {
            "commercial_contact": "",
            "commercial_email": "",
            "commercial_phone": "",
            "technical_contact": "",
            "technical_email": "",
            "technical_phone": "",
            "verticals": "",
        }
    entry = contacts[provider]
    if tech_name and not entry["technical_contact"]:
        entry["technical_contact"] = tech_name
    if tech_email and not entry["technical_email"]:
        entry["technical_email"] = tech_email

wb2.close()
print(f"  Total partners after technical contacts: {len(contacts)}")

# ---------------------------------------------------------------------------
# 3. Read Source of Truth - Features/Verticals
# ---------------------------------------------------------------------------
print("Reading Source of Truth (features)...")

wb3 = openpyxl.load_workbook(SOT_FILE, read_only=True, data_only=True)
ws_partners = wb3["Partners"]
rows_sot = list(ws_partners.iter_rows(values_only=True))
header_sot = [clean(h) for h in rows_sot[0]]

col_prov = header_sot.index("PROVIDER_NAME")
feature_indices = {}
for fc in FEATURE_COLS:
    if fc in header_sot:
        feature_indices[fc] = header_sot.index(fc)

# Aggregate: for each provider, if ANY row has True/1 for a feature, mark it
provider_features = {}  # provider_upper -> set of feature col names

for row in rows_sot[1:]:
    prov = norm_partner(row[col_prov])
    if not prov:
        continue
    if prov not in provider_features:
        provider_features[prov] = set()
    for fc, idx in feature_indices.items():
        if is_truthy(row[idx]):
            provider_features[prov].add(fc)

wb3.close()

# Merge verticals into contacts
for prov, feats in provider_features.items():
    if prov not in contacts:
        contacts[prov] = {
            "commercial_contact": "",
            "commercial_email": "",
            "commercial_phone": "",
            "technical_contact": "",
            "technical_email": "",
            "technical_phone": "",
            "verticals": "",
        }
    labels = [FEATURE_LABELS[f] for f in FEATURE_COLS if f in feats]
    contacts[prov]["verticals"] = ", ".join(labels)

print(f"  Feature data for {len(provider_features)} providers.")
print(f"  Total partners in final dict: {len(contacts)}")

# ---------------------------------------------------------------------------
# 4. Write output
# ---------------------------------------------------------------------------
os.makedirs(OUTPUT_DIR, exist_ok=True)

# Sort partners alphabetically
sorted_partners = sorted(contacts.keys())

lines = ['"""Auto-generated partner contacts dictionary."""\n\n']
lines.append("PARTNER_CONTACTS = {\n")

for partner in sorted_partners:
    data = contacts[partner]
    lines.append(f'    "{escape_str(partner)}": {{\n')
    lines.append(f'        "commercial_contact": "{escape_str(data["commercial_contact"])}",\n')
    lines.append(f'        "commercial_email": "{escape_str(data["commercial_email"])}",\n')
    lines.append(f'        "commercial_phone": "{escape_str(data["commercial_phone"])}",\n')
    lines.append(f'        "technical_contact": "{escape_str(data["technical_contact"])}",\n')
    lines.append(f'        "technical_email": "{escape_str(data["technical_email"])}",\n')
    lines.append(f'        "technical_phone": "{escape_str(data["technical_phone"])}",\n')
    lines.append(f'        "verticals": "{escape_str(data["verticals"])}",\n')
    lines.append("    },\n")

lines.append("}\n")

with open(OUTPUT_FILE, "w", encoding="utf-8") as f:
    f.writelines(lines)

print(f"\nOutput written to: {OUTPUT_FILE}")
print(f"Total partners: {len(sorted_partners)}")
